# ============================================================
# export_transition_excel – BẢN KHÔNG GRIDLINES + KHÔNG ABSORB + INDEX XÁM
# ============================================================

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

HEADER_COLOR = "FFC000"   # vàng đậm
HEADER_COL_GRAY = "D9D9D9"  # xám nhạt
GAP_COLS = 2              # số cột trống giữa các block


def export_transition_excel(matrices_by_mob: dict, parent_fallback: dict, out_path: str):

    thin = Side(border_style="thin", color="000000")
    out_path = Path(out_path).resolve()   # ⭐ lấy full absolute path

    wb = Workbook()
    wb.remove(wb.active)

    for prod, mob_dict in matrices_by_mob.items():

        ws = wb.create_sheet(str(prod)[:31])

        # ===== BỎ GRIDLINES =====
        ws.sheet_view.showGridLines = False

        base_row = 1

        # Tập score & MOB
        all_scores = sorted({str(s) for mob in mob_dict.values() for s in mob.keys()})
        all_mobs = sorted(mob_dict.keys())

        if not all_scores or not all_mobs:
            continue

        # ======================================================
        #  LẶP THEO SCORE
        # ======================================================
        for score in all_scores:

            fb_key = (str(prod), str(score))
            if fb_key not in parent_fallback:
                print(f"⚠️ Không thấy parent fallback {fb_key}")
                continue

            parent_df = parent_fallback[fb_key].round(4).reset_index()
            parent_width = parent_df.shape[1]
            parent_height = parent_df.shape[0] + 2

            block_width = parent_width

            # ==================================================
            #  HEADER SCORE
            # ==================================================
            total_cols = parent_width + GAP_COLS + len(all_mobs) * (block_width + GAP_COLS)

            ws.merge_cells(start_row=base_row, start_column=1,
                           end_row=base_row, end_column=total_cols)
            scell = ws.cell(row=base_row, column=1, value=f"PRODUCT={prod} | SCORE={score}")
            scell.font = Font(bold=True, size=14)
            scell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
            scell.alignment = Alignment(horizontal="left")

            base_row += 1

            # ==================================================
            #  BLOCK PARENT
            # ==================================================
            ws.cell(row=base_row, column=1, value="[PARENT FALLBACK]").font = Font(bold=True)

            for i, row in enumerate(dataframe_to_rows(parent_df, index=False, header=True)):
                for j, val in enumerate(row):

                    c = ws.cell(row=base_row + 1 + i, column=1 + j, value=val)
                    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                    # Header xám
                    if i == 0:
                        c.fill = PatternFill("solid", fgColor=HEADER_COL_GRAY)
                        c.font = Font(bold=True)

                    # CỘT INDEX cũng tô xám
                    if j == 0:
                        c.fill = PatternFill("solid", fgColor=HEADER_COL_GRAY)
                        c.font = Font(bold=True)

                    # Format %
                    if i > 0 and j > 0 and isinstance(val, (int, float)):
                        c.number_format = "0.00%"

            parent_block_height = parent_height

            # ==================================================
            #  HÀM TÍNH CỘT MOB
            # ==================================================
            def mob_col_start(m_idx: int) -> int:
                return 1 + parent_width + GAP_COLS + m_idx * (block_width + GAP_COLS)

            max_mob_height = 0

            # ==================================================
            #  GHI MOB
            # ==================================================
            for mob_idx, mob in enumerate(all_mobs):

                if mob not in mob_dict or score not in mob_dict[mob]:
                    continue

                obj = mob_dict[mob][score]
                P = obj["P"]
                is_fb = obj["is_fallback"]
                reason = obj["reason"]

                df_mat = P.round(4).reset_index()
                n_rows = df_mat.shape[0] + 1
                n_cols = df_mat.shape[1]

                col0 = mob_col_start(mob_idx)

                # Header MOB
                head_text = f"MOB {mob}"
                if is_fb:
                    head_text += " (FALLBACK)"

                ws.merge_cells(start_row=base_row,
                               start_column=col0,
                               end_row=base_row,
                               end_column=col0 + n_cols - 1)
                hcell = ws.cell(row=base_row, column=col0, value=head_text)
                hcell.font = Font(bold=True)
                hcell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
                hcell.alignment = Alignment(horizontal="center")

                data_offset = 1
                if is_fb and reason:
                    rcell = ws.cell(row=base_row + 1, column=col0, value=f"Reason: {reason}")
                    rcell.font = Font(italic=True, size=8)
                    data_offset = 2

                # Ghi dataframe MOB
                for i, row in enumerate(dataframe_to_rows(df_mat, index=False, header=True)):
                    for j, val in enumerate(row):

                        c = ws.cell(row=base_row + data_offset + i, column=col0 + j, value=val)
                        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                        # Header xám
                        if i == 0:
                            c.fill = PatternFill("solid", fgColor=HEADER_COL_GRAY)
                            c.font = Font(bold=True)

                        # CỘT INDEX xám
                        if j == 0:
                            c.fill = PatternFill("solid", fgColor=HEADER_COL_GRAY)
                            c.font = Font(bold=True)

                        # Format %
                        if i > 0 and j > 0 and isinstance(val, (int, float)):
                            c.number_format = "0.00%"

                max_mob_height = max(max_mob_height, data_offset + n_rows)

            # Xuống block mới
            block_height = max(parent_block_height, max_mob_height)
            base_row += block_height + 3

        # ==================================================
        #  Auto column width
        # ==================================================
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    wb.save(out_path)
    print(f"✅ Xuất file transition thành công → {out_path}")
