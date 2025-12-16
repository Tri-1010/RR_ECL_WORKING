# ============================================================
# export_transition_excel – BẢN HOÀN CHỈNH CHO TRANSITION MỚI
# ============================================================

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

HEADER_COLOR = "FFC000"   # vàng đậm
ABSORB_COLOR = "FFF2CC"   # vàng nhạt
GAP_COLS = 2              # số cột trống giữa các block


def export_transition_excel(matrices_by_mob: dict, parent_fallback: dict, out_path: str):

    thin = Side(border_style="thin", color="000000")
    out_path = Path(out_path)

    wb = Workbook()
    wb.remove(wb.active)

    for prod, mob_dict in matrices_by_mob.items():

        ws = wb.create_sheet(str(prod)[:31])
        base_row = 1

        # Tập score & MOB
        all_scores = sorted({str(s) for mob in mob_dict.values() for s in mob.keys()})
        all_mobs = sorted(mob_dict.keys())

        if not all_scores or not all_mobs:
            continue

        # ======================================================
        #  LẶP THEO SCORE (dọc)
        # ======================================================
        for score in all_scores:

            fb_key = (str(prod), str(score))
            if fb_key not in parent_fallback:
                print(f"⚠️ Không thấy parent fallback {fb_key}")
                continue

            # Parent fallback
            parent_df = parent_fallback[fb_key].round(4).reset_index()
            parent_width = parent_df.shape[1]
            parent_height = parent_df.shape[0] + 2  # header + df

            block_width = parent_width  # để căn MOB thẳng hàng với parent

            # ==================================================
            #  HEADER SCORE
            # ==================================================
            total_cols = parent_width + GAP_COLS + len(all_mobs) * (block_width + GAP_COLS)

            ws.merge_cells(start_row=base_row, start_column=1,
                           end_row=base_row, end_column=total_cols)
            scell = ws.cell(row=base_row, column=1, value=f"PRODUCT={prod} | SCORE={score}")
            scell.font = Font(bold=True, size=14)
            scell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
            scell.alignment = Alignment(horizontal="left")

            base_row += 1

            # ==================================================
            #  BLOCK PARENT – LUÔN NẰM BÊN TRÁI
            # ==================================================
            ws.cell(row=base_row, column=1, value="[PARENT FALLBACK]").font = Font(bold=True)

            # ghi df
            for i, row in enumerate(dataframe_to_rows(parent_df, index=False, header=True)):
                for j, val in enumerate(row):
                    c = ws.cell(row=base_row + 1 + i, column=1 + j, value=val)
                    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            parent_block_height = parent_height

            # ==================================================
            # Hàm tính vị trí cột bắt đầu của MOB
            # ==================================================
            def mob_col_start(m_idx: int) -> int:
                return 1 + parent_width + GAP_COLS + m_idx * (block_width + GAP_COLS)

            max_mob_height = 0

            # ==================================================
            #  GHI CÁC MOB NẰM NGANG BÊN PHẢI
            # ==================================================
            for mob_idx, mob in enumerate(all_mobs):

                if mob not in mob_dict or score not in mob_dict[mob]:
                    continue

                obj = mob_dict[mob][score]
                P = obj["P"]
                is_fb = obj["is_fallback"]
                reason = obj["reason"]

                df_mat = P.round(4).reset_index()
                n_rows = df_mat.shape[0] + 1
                n_cols = df_mat.shape[1]

                col0 = mob_col_start(mob_idx)

                # ===== HEADER MOB =====
                head_text = f"MOB {mob}"
                if is_fb:
                    head_text += " (FALLBACK)"

                ws.merge_cells(start_row=base_row,
                               start_column=col0,
                               end_row=base_row,
                               end_column=col0 + n_cols - 1)
                hcell = ws.cell(row=base_row, column=col0, value=head_text)
                hcell.font = Font(bold=True)
                hcell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
                hcell.alignment = Alignment(horizontal="center")

                # Lý do fallback (nếu có)
                data_offset = 1
                if is_fb and reason:
                    rcell = ws.cell(row=base_row + 1, column=col0, value=f"Reason: {reason}")
                    rcell.font = Font(italic=True, size=8)
                    data_offset = 2

                # ===== GHI DATAFRAME MOB =====
                for i, row in enumerate(dataframe_to_rows(df_mat, index=False, header=True)):
                    for j, val in enumerate(row):
                        c = ws.cell(row=base_row + data_offset + i, column=col0 + j, value=val)
                        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # ===== HEATMAP =====
                col_start = col0 + 1
                col_end = col0 + n_cols - 1
                row_start = base_row + data_offset + 1
                row_end = base_row + data_offset + df_mat.shape[0]

                ws.conditional_formatting.add(
                    f"{get_column_letter(col_start)}{row_start}:"
                    f"{get_column_letter(col_end)}{row_end}",
                    ColorScaleRule(
                        start_type="percentile", start_value=10, start_color="63BE7B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="percentile", end_value=90, end_color="F8696B"
                    )
                )

                # ===== Absorbing highlight =====
                ABS = {"PREPAY", "WRITEOFF", "SOLDOUT"}
                for i in range(df_mat.shape[0]):
                    if df_mat.iloc[i, 0] in ABS:
                        for j in range(n_cols):
                            c = ws.cell(row=base_row + data_offset + 1 + i, column=col0 + j)
                            c.fill = PatternFill("solid", fgColor=ABSORB_COLOR)
                            c.font = Font(bold=True)

                max_mob_height = max(max_mob_height, data_offset + n_rows)

            # ==================================================
            #  hết block SCORE → nhảy xuống dưới
            # ==================================================
            block_height = max(parent_block_height, max_mob_height)
            base_row += block_height + 3

        # ==================================================
        #  Auto column width
        # ==================================================
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    wb.save(out_path)
    print(f"✅ Xuất file transition thành công → {out_path}")
