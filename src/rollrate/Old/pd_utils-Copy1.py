# ============================================================
# export_transition_excel – bản hoàn chỉnh có format đẹp
# ============================================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path

ABSORB_COLOR = "FFF2CC"   # vàng nhạt
HEADER_COLOR = "FFC000"   # vàng đậm


def export_transition_excel(matrices_by_mob, out_path: str):
    """
    Xuất ma trận Markov với format đẹp:
    - Mỗi PRODUCT = 1 sheet
    - Score theo chiều dọc
    - MOB theo chiều ngang
    - Merge cell cho block Score + block MOB
    - Heatmap 3 màu trong Excel
    - Absorbing states highlight màu vàng
    """

    wb = Workbook()
    wb.remove(wb.active)
    out_path = Path(out_path)

    thin = Side(border_style="thin", color="000000")

    for prod, mob_dict in matrices_by_mob.items():

        ws = wb.create_sheet(str(prod)[:31])
        base_row = 1

        # Lấy danh sách toàn bộ MOB & Score trong sheet
        all_mobs = sorted(mob_dict.keys())
        all_scores = sorted({s for mob in mob_dict.values() for s in mob.keys()})

        # ===========================
        #  Lặp qua từng SCORE (dọc)
        # ===========================
        for score in all_scores:

            col_cursor = 1

            # --- SCORE header (merge row) ---
            ws.merge_cells(start_row=base_row, start_column=1,
                           end_row=base_row, end_column=1)
            cell = ws.cell(row=base_row, column=1, value=f"Score = {score}")
            cell.font = Font(bold=True, color="000000", size=12)
            cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)

            # ===========================
            #  Lặp theo MOB (ngang)
            # ===========================
            for mob in all_mobs:
                if mob not in mob_dict or score not in mob_dict[mob]:
                    continue

                mat = mob_dict[mob][score]
                df_mat = mat.round(4).reset_index()

                n_rows = df_mat.shape[0] + 2
                n_cols = df_mat.shape[1]

                # --- MOB block header (merge) ---
                ws.merge_cells(start_row=base_row + 1, start_column=col_cursor,
                               end_row=base_row + 1, end_column=col_cursor + n_cols - 1)

                hcell = ws.cell(row=base_row + 1, column=col_cursor,
                                value=f"MOB {mob}")
                hcell.font = Font(bold=True, color="000000")
                hcell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
                hcell.alignment = Alignment(horizontal="center")

                # --- Ghi dataframe ---
                for i, row in enumerate(dataframe_to_rows(df_mat, index=False, header=True)):
                    for j, val in enumerate(row):
                        c = ws.cell(row=base_row + 2 + i, column=col_cursor + j, value=val)
                        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # --- Conditional formatting heatmap ---
                col_start = col_cursor + 1  # bỏ cột state
                col_end = col_cursor + n_cols - 1
                row_start = base_row + 3
                row_end = base_row + 2 + df_mat.shape[0]

                ws.conditional_formatting.add(
                    f"{get_column_letter(col_start)}{row_start}:"
                    f"{get_column_letter(col_end)}{row_end}",
                    ColorScaleRule(
                        start_type="percentile", start_value=10, start_color="63BE7B",  # xanh
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",     # vàng
                        end_type="percentile", end_value=90, end_color="F8696B"      # đỏ
                    )
                )

                # --- Highlight absorbing states ---
                for i in range(df_mat.shape[0]):
                    st = df_mat.iloc[i, 0]
                    if st in ["PREPAY", "WRITEOFF", "SOLDOUT"]:
                        for j in range(n_cols):
                            c = ws.cell(row=base_row + 2 + i, column=col_cursor + j)
                            c.fill = PatternFill("solid", fgColor=ABSORB_COLOR)
                            c.font = Font(bold=True)

                # chuyển sang MOB tiếp theo (n_cols + 2 khoảng trống)
                col_cursor += n_cols + 2

            # block SCORE hoàn thành → xuống block tiếp theo
            base_row += n_rows + 3

        # Auto column width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(out_path)
    print(f"✅ Xuất file ma trận Markov đẹp hoàn chỉnh → {out_path}")
