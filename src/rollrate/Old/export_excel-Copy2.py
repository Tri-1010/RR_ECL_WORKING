# ============================================================
# export_transition_excel – PARENT bên trái, MOB căn thẳng hàng
# ============================================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# Sửa thành 2 dòng này
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path

HEADER_COLOR = "FFC000"   # vàng đậm cho header
ABSORB_COLOR = "FFF2CC"   # vàng nhạt cho absorbing state
GAP_COLS = 2              # số cột trống giữa các block


def export_transition_excel(matrices_by_mob, parent_fallback, out_path: str):
    """
    Xuất ma trận transition với layout:

    Sheet = PRODUCT

    PRODUCT=SALPIL | SCORE=A
        [PARENT FALLBACK]   MOB0   MOB1   MOB2   ...

    PRODUCT=SALPIL | SCORE=B
        [PARENT FALLBACK]   MOB0   MOB1   ...

    - Parent luôn nằm bên trái (cột 1).
    - Các MOB căn ngang bên phải, block-width cố định (dựa trên parent).
    - Score chạy theo chiều dọc, mỗi score là một block.
    """

    thin = Side(border_style="thin", color="000000")
    wb = Workbook()
    wb.remove(wb.active)
    out_path = Path(out_path)

    for prod, mob_dict in matrices_by_mob.items():
        # Tạo sheet theo product
        ws = wb.create_sheet(str(prod)[:31])
        base_row = 1  # dòng bắt đầu của block SCORE hiện tại

        # Lấy danh sách score & mob
        all_scores = sorted({
            str(s) for mob in mob_dict.values() for s in mob.keys()
        })
        all_mobs = sorted(mob_dict.keys())

        # Nếu không có score hoặc mob thì bỏ qua
        if not all_scores or not all_mobs:
            continue

        # ===========================
        #  Lặp theo từng SCORE (dọc)
        # ===========================
        for score in all_scores:
            fb_key = (str(prod), str(score))
            if fb_key not in parent_fallback:
                print(f"⚠️ Không tìm thấy parent_fallback cho (product={prod}, score={score}) → bỏ qua block này.")
                continue

            # --- Chuẩn bị parent matrix & kích thước block chuẩn ---
            parent_df = parent_fallback[fb_key].round(4).reset_index()
            parent_width = parent_df.shape[1]      # số cột parent
            parent_height = parent_df.shape[0] + 2 # 1 dòng label + n dòng dataframe + 1 dòng header DF

            block_width = parent_width             # giả định MOB matrix cùng số cột
            gap = GAP_COLS

            # ----- HEADER lớn: PRODUCT | SCORE -----
            # Ước tính cột cuối cùng của block (parent + tất cả MOB)
            total_blocks = 1 + len(all_mobs)  # 1 parent + N MOB
            last_col = 1 + parent_width + gap + (total_blocks - 1) * (block_width + gap)

            ws.merge_cells(start_row=base_row, start_column=1,
                           end_row=base_row, end_column=last_col)
            big_header = ws.cell(row=base_row, column=1,
                                 value=f"PRODUCT={prod} | SCORE={score}")
            big_header.font = Font(bold=True, size=13)
            big_header.alignment = Alignment(horizontal="left")
            big_header.fill = PatternFill("solid", fgColor=HEADER_COLOR)

            base_row += 1  # dòng dưới big header

            # ===============================
            # 1) GHI PARENT FALLBACK BÊN TRÁI
            # ===============================
            parent_start_col = 1

            # Label "PARENT FALLBACK"
            pf_cell = ws.cell(row=base_row, column=parent_start_col,
                              value="[PARENT FALLBACK]")
            pf_cell.font = Font(bold=True)
            pf_cell.alignment = Alignment(horizontal="left")

            # Ghi parent matrix (bắt đầu từ base_row+1)
            for i, row in enumerate(dataframe_to_rows(parent_df, index=False, header=True)):
                for j, val in enumerate(row):
                    c = ws.cell(row=base_row + 1 + i, column=parent_start_col + j, value=val)
                    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Chiều cao block parent (để tính base_row cho score tiếp theo)
            parent_block_height = parent_height

            # Hàm xác định cột bắt đầu cho từng MOB (căn thẳng hàng theo parent)
            def mob_col_start(mob_index: int) -> int:
                # parent chiếm [1 .. parent_width]
                # khoảng trống gap
                # mỗi MOB chiếm block_width + gap
                return 1 + parent_width + gap + mob_index * (block_width + gap)

            # ===============================
            # 2) GHI CÁC MOB NẰM NGANG BÊN PHẢI
            # ===============================
            max_mob_block_height = 0

            for mob_idx, mob in enumerate(all_mobs):
                if mob not in mob_dict or score not in mob_dict[mob]:
                    continue

                obj = mob_dict[mob][score]
                mat = obj["P"]
                is_fb = obj.get("is_fallback", False)
                reason = obj.get("reason", "")

                df_mat = mat.round(4).reset_index()
                n_rows = df_mat.shape[0] + 1  # 1 dòng header DF + data
                n_cols = df_mat.shape[1]

                # Nếu số cột MOB != parent_width thì vẫn dùng n_cols cho block, nhưng alignment vẫn dựa trên block_width
                col_cursor = mob_col_start(mob_idx)

                # ----- HEADER cho MOB -----
                header_text = f"MOB {mob}"
                if is_fb:
                    header_text += " (FALLBACK)"

                # Merge header ngang trên block MOB
                ws.merge_cells(start_row=base_row,
                               start_column=col_cursor,
                               end_row=base_row,
                               end_column=col_cursor + n_cols - 1)
                hcell = ws.cell(row=base_row, column=col_cursor, value=header_text)
                hcell.font = Font(bold=True)
                hcell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
                hcell.alignment = Alignment(horizontal="center")

                # Nếu muốn in lý do fallback, có thể thêm 1 dòng nhỏ bên dưới header:
                if is_fb and reason:
                    reason_cell = ws.cell(row=base_row + 1, column=col_cursor,
                                          value=f"Reason: {reason}")
                    reason_cell.font = Font(italic=True, size=8)
                    # dữ liệu DF sẽ bắt đầu từ base_row+2 trong trường hợp này
                    data_row_offset = 2
                else:
                    data_row_offset = 1

                # ----- Ghi ma trận MOB -----
                for i, row in enumerate(dataframe_to_rows(df_mat, index=False, header=True)):
                    for j, val in enumerate(row):
                        c = ws.cell(row=base_row + data_row_offset + i,
                                    column=col_cursor + j,
                                    value=val)
                        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # ----- Heatmap cho phần xác suất (bỏ cột STATE) -----
                col_start = col_cursor + 1
                col_end = col_cursor + n_cols - 1
                row_start = base_row + data_row_offset + 1  # bỏ hàng header df
                row_end = base_row + data_row_offset + df_mat.shape[0]

                ws.conditional_formatting.add(
                    f"{get_column_letter(col_start)}{row_start}:"
                    f"{get_column_letter(col_end)}{row_end}",
                    ColorScaleRule(
                        start_type="percentile", start_value=10, start_color="63BE7B",  # xanh
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",       # vàng
                        end_type="percentile", end_value=90, end_color="F8696B"        # đỏ
                    )
                )

                # ----- Highlight absorbing states -----
                absorbing_states = ["PREPAY", "WRITEOFF", "SOLDOUT"]
                for i in range(df_mat.shape[0]):
                    st = df_mat.iloc[i, 0]
                    if st in absorbing_states:
                        for j in range(n_cols):
                            c = ws.cell(row=base_row + data_row_offset + 1 + i,
                                        column=col_cursor + j)
                            c.fill = PatternFill("solid", fgColor=ABSORB_COLOR)
                            c.font = Font(bold=True)

                # Cập nhật chiều cao block MOB lớn nhất
                mob_block_height = data_row_offset + n_rows
                max_mob_block_height = max(max_mob_block_height, mob_block_height)

            # ===============================
            # 3) XONG 1 BLOCK SCORE → NHẢY XUỐNG DƯỚI
            # ===============================
            block_height = max(parent_block_height, max_mob_block_height)
            base_row += block_height + 3  # +3 dòng trống giữa các score

        # Auto-fit chiều rộng cột
        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    wb.save(out_path)
    print(f"✅ Xuất Excel thành công → {out_path}")
